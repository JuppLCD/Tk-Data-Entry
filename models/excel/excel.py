import os

from openpyxl import Workbook, load_workbook


class ExcelModel():
    def __init__(self,  heading: tuple, sheet_name: str | None):
        self.heading = heading
        self.sheet_name = sheet_name

        ruta_actual = os.path.abspath(os.path.dirname(__file__))
        self.filepath = os.path.join(
            ruta_actual, '..', '..', 'db', 'data.xlsx'
        )

        self.openWB()
        self.save_excel()

    def apend_data(self, all_data: list[tuple]):
        self.openWB()
        for data in all_data:
            self.sheet.append(data)
        self.save_excel()

    def openWB(self):
        if os.path.exists(self.filepath):
            self.workbook = load_workbook(self.filepath)
            self.get_sheet()
        else:
            self.workbook = Workbook()

            self.get_sheet()
            self.sheet.append(self.heading)

    def get_sheet(self):
        if self.sheet_name == None:
            self.sheet = self.workbook.active
        else:
            if not self.sheet_name in self.workbook.sheetnames:
                self.workbook.create_sheet(self.sheet_name)

            self.sheet = self.workbook[self.sheet_name]

    def save_excel(self):
        self.workbook.save(self.filepath)
