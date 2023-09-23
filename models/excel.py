import os

from openpyxl import Workbook, load_workbook

from utils.entities import User
from models.model_abstract import UserModelAbstract

ruta_actual = os.path.abspath(os.path.dirname(__file__))
filepath = os.path.join(ruta_actual, '..', 'db', 'data.xlsx')


class UserModel(UserModelAbstract):
    def __init__(self):
        if not os.path.exists(filepath):
            self.workbook = Workbook()
            self.sheet = self.workbook.active

            heading = ["First Name", "Last Name", "Title", "Age",
                       "Nationality", "# Courses", "# Semesters", "Registration status"]
            self.sheet.append(heading)
        else:
            self.workbook = load_workbook(filepath)
            self.sheet = self.workbook.active

        self.save_excel()

    def store(self, data: User):
        self.openWB()

        self.sheet.append([
            data.firstname,
            data.lastname,
            data.title,
            data.age,
            data.nationality,
            data.numcourses,
            data.numsemesters,
            data.registration_status
        ]
        )

        self.save_excel()

    def openWB(self):
        self.workbook = load_workbook(filepath)
        self.sheet = self.workbook.active

    def save_excel(self):
        self.workbook.save(filepath)
